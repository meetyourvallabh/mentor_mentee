from flask import Flask, render_template, request, flash, redirect, url_for, session, json, jsonify, make_response, send_from_directory, send_file
import os
import requests
from flask_pymongo import PyMongo
from flask_mail import Mail, Message
from random import *
import bcrypt
from functools import wraps
from datetime import datetime
import pdfkit
import itertools
import socket
from openpyxl import Workbook

basedir = os.path.abspath(os.path.dirname(__file__))
config = pdfkit.configuration(wkhtmltopdf="C:\Program Files\wkhtmltopdf\\bin\wkhtmltopdf.exe")

app = Flask(__name__)

app.config.from_envvar('APPLICATION_SETTINGS')


ALLOWED_EXTENSIONS = set(['pdf', 'docx', 'doc'])
UPLOAD_DIRECTORY = "/static/"

app.config['MONGO_DBNAME'] = 'mm'
app.config['MONGO_URI'] = 'mongodb://127.0.0.1:27017/mm'
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'developer@makeyourown.club'
app.config['MAIL_PASSWORD'] = 'myocoo@123'


mongo = PyMongo(app)
mail = Mail(app)

ip = socket.gethostbyname(socket.gethostname())


@app.context_processor
def for_whole_application():
    branches = mongo.db.branches
    all_branches = branches.find()

    user1 = mongo.db.users
    reg_mentors = user1.find({"type":"mentor"})

    result = [all_branches, reg_mentors]
    return dict(branches = result)

# Check if user logged in
def is_logged_in(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'logged_in' in session:
            return f(*args, **kwargs)
        else:
            flash('Please Login First', 'danger')
            return redirect(url_for('login'))
    return wrap

# Check if user is admin
def is_admin(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if session['type'] == 'admin':
            return f(*args, **kwargs)
        else:
            flash('Unauthorized access, You\'re not allowed here', 'danger')
            return redirect(url_for('index'))
    return wrap

# Check if user is hod
def is_hod(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if session['type'] == 'admin' or session['type'] == 'hod':
            return f(*args, **kwargs)
        else:
            flash('Unauthorized access, You\'re not allowed here', 'danger')
            return redirect(url_for('index'))
    return wrap

# Check if user is mentor
def is_mentor(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if session['type'] == 'admin' or session['type'] == 'mentor' or session['type'] == 'hod':
            return f(*args, **kwargs)
        else:
            flash('Unauthorized access, You\'re not allowed here', 'danger')
            return redirect(url_for('index'))
    return wrap
# Check if user is hod
def is_only_mentee(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if session['type'] == 'mentee':
            return f(*args, **kwargs)
        else:
            flash('Unauthorized access, You\'re not allowed here', 'danger')
            return redirect(url_for('index'))
    return wrap


@app.route('/')
def index():
    date = datetime.today()
    users = mongo.db.users
    meetings = mongo.db.meetings
    meeting_requests = []
    if 'logged_in' in session:
        if session['type'] == 'mentee':
            mentee = users.find_one({'email':session['email']})
            mentor = users.find_one({'email':mentee['mentor_email']})
            if mentor is not None:
                all_meetings = meetings.find({'mentor_email':mentor['email']})
                
                if all_meetings is not None:
                    for r in all_meetings:
                        meetdate = datetime.strptime(r['date'], '%Y-%m-%d')
                        print(meetdate)
                        if meetdate >= date:
                            meeting_requests.append({'venue':r['venue'],'time':r['time'],'date':r['date'],'subject':r['subject']})
            return render_template("index.html",meeting_requests =meeting_requests)

        if session['type'] == 'mentor':
            meeting_requests1 = []
            mentor1 = users.find_one({'email':session['email']})
            all_meetings = meetings.find({'mentor_email':session['email']})
            for r in all_meetings:
                meetdate = datetime.strptime(r['date'], '%Y-%m-%d')
                print(meetdate)
                if meetdate >= date:
                    meeting_requests1.append({'venue':r['venue'],'time':r['time'],'date':r['date'],'subject':r['subject']})
            print(meeting_requests1)
            return render_template("index.html",meeting_requests1 = meeting_requests1,mentor1=mentor1)
    



    return render_template("index.html")

@app.route('/login',methods = ['POST','GET'])
def login():
    if request.method == 'POST':
        
        email = request.form['email']
        password = request.form['password']
        users = mongo.db.users
        founduser = users.find_one({'email':email})
        if founduser:
            if bcrypt.checkpw(password.encode('utf-8'), founduser['password']):
                print('found user')
                if founduser['type'] == 'mentee':
                    if founduser['verification'] == 'yes':
                        session['fname'] = founduser['fname']
                        session['mname'] = founduser['mname']
                        session['lname'] = founduser['lname']
                        session['type'] = founduser['type']
                        session['email']=founduser['email']
                        session['year']=founduser['year']
                        session['branch']=founduser['branch']
                        session['mentor']=founduser['mentor_email']
                        session['mentor_name']=founduser['mentor']
                        session['logged_in']=True
                        
                        session['branch'] = founduser['branch']
                        flash('Login Successfull','success')
                        return redirect(url_for('index'))
                    else:
                        flash('Your Email is not verified yet, Please Do email verification by using link provided in mail','secondary')
                        flash('Wait for the email if not recieved','warning')
                        return redirect(url_for('login'))
                else:
                    session['fname'] = founduser['fname']
                    session['mname'] = founduser['mname']
                    session['lname'] = founduser['lname']
                    session['type'] = founduser['type']
                    session['email']=founduser['email']
                    session['logged_in']=True
                    if founduser['type'] in ['hod','mentor']:
                            session['branch'] = founduser['branch']
                    flash('Login Successfull','success')
                    return redirect(url_for('index'))
            else:
                flash('Wrong Email ID and Password','danger')
    
    return render_template("login.html")


@app.route('/changepassword',methods = ['POST','GET'])
def changepassword():
    users = mongo.db.users
    if request.method == 'POST':
        founduser = users.find_one({'email':request.form['email']})
        if founduser:
            token=str(randint(11111,99999))
            msg = Message('MM: Password reset link', sender='makeyourown48@gmail.com', recipients=[request.form['email']])
            msg_string = '<h1>Hello, ' + founduser['fname'] + '</h1><br><br> Click on below button to change your account password: <br>'+str(ip)+'/changepasswordtoken/'+token
            msg.body = msg_string
            msg.html = msg.body
            mail.send(msg)
            users.update({'email':request.form['email']},{'$set':{'resetlink':token}},upsert = True)
            flash('Password reset link sent successfully','success')
        else:
            flash('No user acount found for this email','danger')
    return redirect(url_for('login'))


@app.route('/changepasswordtoken/<token>',methods = ['POST','GET'])
def changepasswordtoken(token):
    users = mongo.db.users
    if request.method == 'POST':
        hashpass = bcrypt.hashpw(request.form['password'].encode('utf-8'), bcrypt.gensalt())
        done = users.update({'resetlink':token},{'$set':{'password':hashpass,'verification':'yes'}})
        if done:
            users.update({'resetlink':token},{'$set':{'resetlink':'inactive'}})
            flash('Password changed successfully','success')
        else:
            flash('Invalid Token','danger')
            return redirect(url_for('login'))
    else:
        validtoken = users.find_one({'resetlink':token})
        if validtoken:
            return render_template("resetpassword.html",token=token)
        else:
            flash('Invalid Token','danger')
            return redirect(url_for('login'))

    return redirect(url_for('login'))

@app.route('/register',methods = ['POST','GET'])
def register():
    users = mongo.db.users
    
    
    if request.method == 'POST':
        founduser = users.find_one({'email':request.form['email']})
        verification_code = str(randint(1111,9999))
        if founduser is None:
            passw = request.form['password']
            hashpass = bcrypt.hashpw(passw.encode('utf-8'), bcrypt.gensalt())
            
            #get mtor email of selected mtor
            mentor_email = request.form.get('mentor')
            get_mentor_name = users.find_one({'email':mentor_email, 'type':'mentor'})
            mentor_name = get_mentor_name['fname']+" "+get_mentor_name['lname']
            print('email ----{}'.format(str(mentor_email)))
            #create mtee user
            users.insert_one({'fname':request.form['fname'],'password':hashpass,'verification':verification_code,'approved':'no','mname':request.form['mname'],'lname':request.form['lname'],'type':'mentee' ,
            'email':request.form['email'],'phone':request.form['phone'], 'branch':request.form['branch'],'year':request.form['year'],'division':request.form['division'],'batch':request.form['batch'],'mentor':mentor_name,'mentor_email':mentor_email, 'status':'Regular'})
            msg = Message('MM: Mentee account registered successfully', sender='makeyourown48@gmail.com', recipients=[request.form['email']])
            msg_string = '<h1>Hello ' + request.form['fname']+ request.form['lname'] + '</h1><br> Account registered sucessfully !! <br><br> Click on below link to verify your account: <br> http://mm.mcoeit.com/verify/'+verification_code+'<br><p style="color:red;">IMP!! If link not working!<br>Verify through college network</p><br>'+str(ip)+'/verify/'+verification_code+'<br>THANK YOU.'
            msg.body = msg_string
            msg.html = msg.body
            mail.send(msg)
            flash('Mentee account registered successfully.','success')
            flash('Verification link sent. Check EMAIL!.','secondary')
            return redirect(url_for('login'))
        else:
            flash('Email id already exists','danger')
    return render_template("register.html")

@app.route('/verify/<code>',methods = ['POST','GET'])
def verify(code):
    if request.method == 'POST' or 'GET':
        users = mongo.db.users
        user = users.find_one({'verification':code})
        if user is None:
            flash('Verification link is expired or already verified','warning')
            return redirect(url_for('login'))
        else:
            users.update({'verification':code},{'$set':{'verification':'yes'}})
            session['email'] = user['email']
            session['fname'] = user['fname']
            session['lname'] = user['lname']
            session['mname'] = user['mname']
            session['year'] = user['year']
            session['branch'] = user['branch']
            session['type'] = user['type']
            session['logged_in']=True
            session['verification'] = 'yes'
            flash('Email verification done successfully','success')
            return redirect(url_for('index'))
    return redirect(url_for('index'))


@app.route('/add_mentor',methods = ['POST','GET'])
@is_logged_in
@is_mentor
def add_mentor():
    users = mongo.db.users
    branches = mongo.db.branches
    all_branches = branches.find()
    user = users.find_one({'email':session['email']})
    if session['type'] == 'hod':
        branch = branches.find_one({'hod_email':session['email']})
        mentors = users.find({'type':'mentor','branch':branch['branch']})
    else:
        mentors = users.find({'type':'mentor'})
        mentors1 = mentors.clone()
        # mentors1 = users.find({'type':'mentor'})
    if request.method == 'POST':
        passw = request.form['fname']+'.'+str(randint(1111,9999))
        hashpass = bcrypt.hashpw(passw.encode('utf-8'), bcrypt.gensalt())
        founduser = users.find_one({'email':request.form['email']})
        # existingbatch = users.find_one({'email':request.form['email'],'batch.batch_name':request.form['batch'],'batch.division':request.form['division'],'batch.branch':request.form['branch'],'batch.year':request.form['year']})
        if founduser is None:
            users.insert_one({'fname':request.form['fname'],'password':hashpass,'mname':request.form['mname'],'lname':request.form['lname'],'type':'mentor' , 
            'email':request.form['email'],'phone':request.form['phone'],'branch': request.form['branch']})
            msg = Message('MM: Successfull Mentor account registration', sender='makeyourown48@gmail.com', recipients=[request.form['email']])
            # msg_string = '<h1>Respected Prof. ' + request.form['fname']+ request.form['lname'] + '</h1><br> You are registered sucessfully by Prof.'+ session['fname'] + session['lname'] +' as a Mentor of Batch '+ request.form['batch'] +' of class '+ request.form['year']+' ' +request.form['branch'] +' ' +request.form['division']+'!! <br><br> You can now login using following credentials<br> Email id : '+request.form['email']+'<br> Password : '+passw
            msg_string = '<h1>Respected Prof. ' + request.form['fname']+ request.form['lname'] + '</h1><br> You are registered sucessfully by Prof.'+ session['fname'] + session['lname'] +' as a Mentor in Information Technology Department!! <br><br> You can now login using following credentials<br> Email id : '+request.form['email']+'<br> Password : '+passw
            msg.body = msg_string
            msg.html = msg.body
            mail.send(msg)
            flash('Mentor Added Successfully','success')
            return redirect(url_for('add_mentor'))
        
        # if existingbatch is None:
        #     users.update({'email' : request.form['email']},{'$push':{'batch':{'batch_name':request.form['batch'], 'branch': request.form['branch'], 'division':request.form['division'],  'year':request.form['year']}}})                
        #     msg = Message('MM: New Batch assigned !!', sender='makeyourown48@gmail.com', recipients=[request.form['email']])
        #     msg_string = '<h1>Respected Prof. ' + request.form['fname']+ request.form['lname'] + '</h1><br> New batch is assigned by Prof.'+ session['fname'] + session['lname'] +' as a Mentor of Batch '+ request.form['batch'] +' of class '+ request.form['year']+' ' +request.form['branch'] +' ' +request.form['division']
        #     msg.body = msg_string
        #     msg.html = msg.body
        #     mail.send(msg)
        #     flash('Mentor account already exist','warning')
        #     flash('Batch Assigned Successfully','success')
        #     return redirect(url_for('add_mentor'))                
        else:
            flash('Mentor account already exist','warning')
            flash('Already assigned this batch','danger')
            return redirect(url_for('add_mentor'))
                                
    mentee_count = {}  
    for men in mentors1:
        _count = users.count_documents({'type':"mentee",'mentor_email':men['email']})
        mentee_count.__setitem__(men['email'], _count) 
    return render_template("add_mentor.html",mentors = mentors,branches = all_branches, mentee_count=mentee_count)

@app.route('/add_hod',methods = ['POST','GET'])
@is_logged_in
@is_admin
def add_hod():
    users = mongo.db.users
    branches = mongo.db.branches
    hods = users.find({'type':'hod'})
    if request.method == 'POST':
        hod_found = users.find_one({'email':request.form['email'], 'type':'hod'})
        if hod_found is None:    
            branch_found = branches.find_one({'branch':request.form['branch']})
            if branch_found is None:
                passw = request.form['fname']+'.'+str(randint(1111,9999))
                hashpass = bcrypt.hashpw(passw.encode('utf-8'), bcrypt.gensalt())
                
                users.insert_one({'fname':request.form['fname'],'password':hashpass,'mname':request.form['mname'],'lname':request.form['lname'],'type':'hod' ,
                'email':request.form['email'],'phone':request.form['phone'], 'branch':request.form['branch']})
                
                branches.insert_one({'branch':request.form['branch'],'hod':request.form['fname']+' '+request.form['mname']+' '+request.form['lname'],'hod_email':request.form['email']})
                
                msg = Message('MM: Successfull HOD account registration', sender='makeyourown48@gmail.com', recipients=[request.form['email']])
                msg_string = '<h1>Respected Prof. ' + request.form['fname']+ request.form['lname'] + '</h1><br> You are registered sucessfully !! <br><br> You can now login using following credentials<br> Email id : '+request.form['email']+'<br> Password : '+passw
                msg.body = msg_string
                msg.html = msg.body
                mail.send(msg)
                flash('HOD Added Successfully','success')
                flash('Branch '+request.form['branch']+' Added Successfully','success')
            else:
                flash('Branch already exists!','danger')
        else:
            flash('HOD already exists!','danger')
    return render_template("add_hod.html",hods = hods)


@app.route('/add_mentee',methods = ['POST','GET'])
@is_logged_in
@is_mentor
def add_mentee():
    users = mongo.db.users
    
    mentor = users.find_one({'email':session['email']})
    mentees = users.find({'type':'mentee','mentor_email':session['email']})
    
    print(mentor['fname'])

    if request.method == 'POST':
        batchname = request.form['batch'].split(" ")
        passw = request.form['fname']+'.'+str(randint(1111,9999))
        hashpass = bcrypt.hashpw(passw.encode('utf-8'), bcrypt.gensalt())
        
        users.insert_one({'fname':request.form['fname'],'password':hashpass,'mname':request.form['mname'],'verification':'yes','approved':'yes','lname':request.form['lname'],'type':'mentee' ,
        'email':request.form['email'],'phone':request.form['phone'], 'mentor_email':session['email']})
        
        
        
        msg = Message('MM: Successfull Mentee account registration', sender='makeyourown48@gmail.com', recipients=[request.form['email']])
        msg_string = '<h1>Hello ' + request.form['fname']+ request.form['lname'] + '</h1><br> You are registered sucessfully !! <br><br> You can now login using following credentials<br> Email id : '+request.form['email']+'<br> Password : '+passw
        msg.body = msg_string
        msg.html = msg.body
        mail.send(msg)
        flash('Mentee Added Successfully','success')
        return redirect(url_for('add_mentee'))
    
    return render_template("add_mentee.html",mentees = mentees,mentor = mentor)



@app.route('/delete_mentee/<email>')
@is_logged_in
@is_mentor
def delete_mentee(email):
    users = mongo.db.users
    users.delete_one({'email':email,'type':'mentee'})
    
    flash('Mentee deleted successfully','danger')
    if session['type'] == 'admin':
        return redirect(url_for('dashboard'))
    return redirect(url_for('add_mentee'))


@app.route('/profile',methods = ['POST','GET'])
@is_logged_in
def profile():
    users = mongo.db.users
    user = users.find_one({'email':session['email']})
    states = mongo.db.states
    state = states.find()
    subjects = mongo.db.subjects
    all_subjects = subjects.find_one({'branch':session['branch']})
    sub_data=[]
    #print(all_subjects['subjects'])
    #substr = '{'
    #{u'semester6': [u'DSBDA':{}, u'SP':{}, u'DAA', u'CC', u'CNT'], u'semester5': [u'HCI', u'OS', u'SEPM', u'TOC', u'DBMS']}
    #for s in all_subjects['subjects']:
    #    substr = substr+ '\''+s+'\':['
    #    for sub in all_subjects['subjects'][s]:
    #        substr = substr+ '\''+sub+'\':{},'
    #    substr = substr[:-1]
    #    substr = substr+ '],'
    #substr = substr[:-1]
    #substr = substr+ '}'
    #   
    #print(substr)
    if 'result' not in user:
        users.update_one({'email':session['email']},{'$set':{'result':[{'semester':1,'status':'null','subject':[]},{'semester':2,'status':'null','subject':[]},{'semester':3,'status':'null','subject':[]},{'semester':4,'status':'null','subject':[]},{'semester':5,'status':'null','subject':[]},{'semester':6,'status':'null','subject':[]},{'semester':7,'status':'null','subject':[]},{'semester':8,'status':'null','subject':[]}]}})
        for s in all_subjects['subjects']:
                        
            for s1 in s['subject']:

                
                users.update_one({'email':session['email'],'result.semester':int(s['semester'])},{'$push':{'result.$.subject':{'subject_name':s1,'status':'null','attempts':1}}})
            

    else:
        for s in all_subjects['subjects']:        
            for s1 in s['subject']:
                
                #print(s['semester']+' '+s1)
                found = users.find_one({'email':session['email'],'result.semester':int(s['semester']),'result.subject.subject_name':s1})
                if found is not None:
                    #print('found')
                    pass
                else:
                    #print('not found')
                #users.update_one({'email':session['email'],'result.semester':int(s['semester'])},{'$pull':{'result.$.subject':{'subject_name':s1}}})
                    users.update_one({'email':session['email'],'result.semester':int(s['semester'])},{'$push':{'result.$.subject':{'subject_name':s1,'status':'null','attempts':1}}})


        

    return render_template("profile.html",user = user,states = state[0]['states'],all_subjects=all_subjects)


@app.route('/profile/<step>',methods = ['POST','GET'])
@is_logged_in
def updateprofile(step):
    session['step'] = 'perosnal'
    if request.method == 'POST':
        users = mongo.db.users
        user = users.find_one({'email':session['email']})
        
        if user['type'] in ['mentor','hod','admin']:
            flash('You can\'t view or complete non mentee account profile','danger')
            return redirect(url_for('index'))
        else:
            if step == '1':
                session['step'] = 'personal'
                done = users.update_one({'email' : request.form['email']},{'$set':{'fname':request.form['fname'],'mname':request.form['mname'],'lname':request.form['lname'],'email':request.form['email'],'dob':request.form['dob'],'phone':request.form['phone'],'prn':request.form['prn'],'admission_year':request.form['admission_year'],'personal.hobbies':request.form['hobbies'],'personal.weakness':request.form['weakness'],'personal.strength':request.form['strength'],'personal.majorillness':request.form['majorillness'],'personal.majoralergic':request.form['majoralergic'],'personal.blood_group':request.form['blood_group'],'personal.caste':request.form['caste'],'personal.religion':request.form['religion'],'admission_category':request.form['admission_category']}})
                if done:
                    flash('Personal information updated successfully !!','success')
                    return redirect(url_for('profile'))
                else:
                    flash('Couldn\'t update profile, Try again !!','danger')
                    return redirect(url_for('profile'))
            
            if step == '2':
                session['step'] = 'family'
                done = users.update_one({'email':session['email']},{'$set':{'family.father.fname':user['mname'],'family.father.lname':user['lname'],'family.father.mname':request.form['father_mname'],
                'family.mother.mname':user['mname'],'family.mother.lname':user['lname'],'family.mother.fname':request.form['mother_fname'],
                'family.father.occupation':request.form['father_occupation'],'family.mother.occupation':request.form['mother_occupation'],
                'family.father.phone':request.form['father_phone'],'family.mother.phone':request.form['mother_phone'],'address.permanent.city':request.form['permanent_city'],
                'address.permanent.state':request.form['permanent_state'],'address.permanent.address':request.form['permanent_address'],
                'guardian.name':request.form['guardian_name'],'guardian.phone':request.form['guardian_phone'],'guardian.address':request.form['guardian_address'],'address.local.city':request.form['local_city'],'address.local.state':request.form['local_state'],'address.local.address':request.form['local_address']}})
                if done:
                    flash('Family details updated successfully !!','success')
                else:
                    flash('Couldn\'t update profile, Try again !!','danger')
            if step == '3':
                session['step'] = 'academics'
                done = users.update_one({'email':session['email']},{'$set':{'academic.result.class_10.board':request.form['board10'],'academic.result.class_10.year':request.form['year10'],'academic.result.class_10.marks':request.form['marks10'],'academic.result.class_10.percentage':request.form['percentage10'],
                'academic.result.class_12.board':request.form['board12'],'academic.result.class_12.year':request.form['year12'],'academic.result.class_12.marks':request.form['marks12'],'academic.result.class_12.percentage':request.form['percentage12'],'academic.result.class_12.marks':request.form['marks12'],'academic.result.class_12.entrance':request.form['entrance12'],
                'academic.result.diploma.board':request.form['board_diploma'],'academic.result.diploma.year':request.form['year_diploma'],'academic.result.diploma.marks':request.form['marks_diploma'],'academic.result.diploma.percentage':request.form['percentage_diploma'],'academic.extra_curicular':request.form['extra_curicular']}})
                if done:
                    flash('Marks details updated successfully !!','success')
                else:
                    flash('Couldn\'t update profile, Try again !!','danger')

            if step == '4':
                session['step'] = 'results'
                subjects = mongo.db.subjects
                all_subjects = subjects.find_one({'branch':session['branch']})
                sub_list=[]
                semester_no=int(request.args.get('semester'))
                #result = users.find_one({'email':session['email'],'result.semester':request.args.get('semester')})
                #print(result)

                #for semester in user['result']:
                #    if semester['semester'] == int(request.args.get('semester')):
                #        for subject in semester['subject']:
                #            print(subject['subject_name'])
                try:
                    sgpa = float(request.form['sgpa'])
                    if not sgpa < 10.1:
                        flash('Please enter number between 1.0 to 10.0 in sgpa field','danger')
                        return redirect(url_for('profile'))

                except:
                    flash('Please enter number in sgpa field','danger')
                    return redirect(url_for('profile'))

                
                for semester in user['result']:
                    sem_status = 'all clear'
                    if semester['semester'] == int(request.args.get('semester')):
                        for subject in semester['subject']:
                            s1 = subject['subject_name']
                            
                            users.update_one({'email':session['email'],'result.semester':semester_no},{'$pull':{'result.$.subject':{'subject_name':s1}}})
                            
                            users.update_one({'email':session['email'],'result.semester':semester_no},{'$push':{'result.$.subject':{'subject_name':s1,'status':request.form[s1+'_status'],'attempts':int(request.form[s1+'_attempt'])}}})
                            if request.form[s1+'_status'] == 'fail':
                                sem_status = 'ATKT'
                                # subject_failed = users.count_documents({"email":session['email'], "result.subject.subject_name":s1, 'result.semester':semester_no+1})
                                subject_failed =  users.count_documents({"result":{"$elemMatch":{"subject.subject_name":s1, "semester":semester_no+1}}})
                                print("s1 ",str(s1))
                                print("semester ",str(semester_no))
                                print("subject_failed === ",str(subject_failed))
                                if subject_failed:
                                    print("subject alrady failed")
                                    # users.update_one({'email':session['email'],'result.semester':semester_no+1},{'$addToSet':{'result.$.subject':{'subject_name':s1,'status':'null','attempts':int(request.form[s1+'_attempt'])+1}}})
                                else:
                                    users.update_one({'email':session['email'],'result.semester':semester_no+1},{'$addToSet':{'result.$.subject':{'subject_name':s1,'status':'null','attempts':int(request.form[s1+'_attempt'])+1}}})
                                    print('updated in '+str(semester_no+1))
                                
                                
                            elif request.form[s1+'_status'] == 'pass':
                                
                                #users.update_one({'email':session['email'],'result.semester':semester_no-1},{'$push':{'result.$.subject':{'subject_name':s1,'status':'pass','attempts':int(request.form[s1+'_attempt'])}}})
                                users.update_one({'email':session['email'],'result.semester':semester_no+1},{'$pull':{'result.$.subject':{'subject_name':s1}}})
                        users.update_one({'email':session['email'],'result.semester':semester_no},{'$set':{'result.$.status':sem_status,'result.$.sgpa':sgpa}})

                flash('Data updated successfully','success')
    return redirect(url_for('profile'))


@app.route('/upload',methods = ['POST','GET'])
@is_logged_in
def upload():
    users = mongo.db.users
    user = users.find_one({'email':session['email']})
    if session['type'] == 'mentee':
        path = os.path.abspath(basedir+'/static/img/'+session['fname']+' '+session['mname']+' '+session['lname']+' '+user['year']+' '+user['branch']+' '+user['division'])
    else:
        path = os.path.abspath(basedir+'/static/img/'+session['fname']+' '+session['mname']+' '+session['lname']+' '+user['branch'])
    if not os.path.exists(path):
        os.makedirs(path)
    app.config['UPLOAD_FOLDER'] = path
    if 'propic' not in request.files:
        flash('No file part')
        return redirect(request.url)
    if 'propic' in user:
        if os.path.exists('static/'+user['propic']):
            os.remove('static/'+user['propic'])
    if request.method == 'POST':
        file = request.files['propic']
        f = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    
        # add your custom code to check that the uploaded file is a valid image and not a malicious file (out-of-scope for this post)
        file.save(f)
        
        users.update_one({'email':session['email']},{'$set':{'propic':'img/'+session['fname']+' '+session['mname']+' '+session['lname']+' '+user['year']+' '+user['branch']+' '+user['division']+'/'+file.filename}})
        flash('Image uploaded successfully','success')
    return redirect(url_for('profile'))

@app.route('/documents',methods = ['POST','GET'])
@is_logged_in
def documents():
    users = mongo.db.users
    user = users.find_one({'email':session['email']})
    if request.method == 'POST':
        path = os.path.abspath(basedir+'/static/img/'+session['fname']+' '+session['mname']+' '+session['lname']+' '+user['year']+' '+user['branch']+' '+user['division']+'/documents')
        print(path)
        if not os.path.exists(path):
            os.makedirs(path)
        app.config['UPLOAD_FOLDER'] = path
        if 'document' not in request.files:
            flash('No file part')
            return redirect(request.url)
        if request.method == 'POST':
            file = request.files['document']
            f = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            
            # add your custom code to check that the uploaded file is a valid image and not a malicious file (out-of-scope for this post)
            file.save(f)
            
            doc_id = str(randint(11,99)) +request.form['caption'][:10]+ str(randint(11,99))
            users.update_one({'email':session['email']},{'$push':{'document':{'doc_id':doc_id,'caption':request.form['caption'],'path':'img/'+session['fname']+' '+session['mname']+' '+session['lname']+' '+user['year']+' '+user['branch']+' '+user['division']+'/documents/'+file.filename}}})
            flash('Image uploaded successfully','success')
            
    return render_template("documents.html",user=user)

@app.route('/delete_document/<id>', methods=['GET', 'POST'])
@is_logged_in
def delete_document(id):
    users = mongo.db.users
    user = users.find_one({'email':session['email']},{'document': {'$elemMatch': {'doc_id':id}}})
    #for doc in user['document']:
    #    print(doc)
    print(user['document'][0]['path'])
    #print(user)
    if os.path.exists(basedir+'/static/'+user['document'][0]['path']):
        os.remove(basedir+'/static/'+user['document'][0]['path'])
    
    users.update_one({'email': session['email']}, 
    { '$pull': { "document" : { 'doc_id': id } } })
    flash('Document deleted successfully','success')

    return redirect(url_for('documents'))


@app.route('/view_documents/<email>')
@is_logged_in
@is_mentor
def view_documents(email):
    users = mongo.db.users
    user = users.find_one({'email':email})
    return render_template("view_documents.html",user = user)


@app.route('/view_profile/<email>',methods = ['POST','GET'])
@is_logged_in
@is_mentor
def view_profile(email):
    users = mongo.db.users
    user = users.find_one({'email':email})
    if user['type'] in ['mentor','hod','admin']:
        flash('You can\'t view non mentee account profile','danger')
        return redirect(url_for('index'))
    return render_template("view_profile.html",user = user)


@app.route('/pdf_profile/<email>/<option>',methods = ['POST','GET'])
@is_logged_in
def pdf_profile(email,option):
    if session['type'] in ['hod','admin','mentor']:
        users = mongo.db.users
        user = users.find_one({'email':email})
        rendered =  render_template("pdf_profile.html",user = user)
        #rendered =  render_template("example.html")
        css = [basedir+'/static/assets/css/view_profile.css',basedir+'/static/assets/bootstrap/css/bootstrap.min.css']
        pdf = pdfkit.from_string(rendered, False, css = css,configuration=config)
        #pdf = pdfkit.from_string(rendered, False, css = css)
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        if option == 'view':
            response.headers['Content-Disposition'] = 'inline; filename=profile.pdf'
        elif option == 'download':
            response.headers['Content-Disposition'] = 'attachment; filename='+user['fname'] +' '+user['mname']+' '+user['lname']+'.pdf'
        else:
            return redirect(url_for('view_profile',email=email))
        return response
  
@app.route('/send_mail',methods = ['POST','GET'])
@is_logged_in
@is_mentor
def send_mail():
    if request.method == 'POST':
        users = mongo.db.users
        all_users = users.find()
        all_users = users.find()
        batch = []
        if session['type'] in ['mentor']:
            user = users.find_one({'email':session['email']})
            # for b in user['batch']:
            #     batch.append(b)
            # bat = batch[0]
            # print(bat['batch_name'])
            # all_mentees = users.find({'type':'mentee','branch':user['branch'],'batch':bat['batch_name'],'division':bat['division'],'year':bat['year']})
            all_mentees = users.find({'type':'mentee','mentor_email':session['email']})
        if session['type'] in ['hod']:
            user = users.find_one({'email':session['email']})
            all_mentors = users.find({'type':'mentor','branch':user['branch']})
            all_mentees = users.find({'type':'mentee','branch':user['branch']})
        if session['type'] in ['admin']:
            all_hods = users.find({'type':'hod'})
            all_mentors = users.find({'type':'mentor'})
            all_mentees = users.find({'type':'mentee'})
        all_receipents = []
        if request.form['receiver'] == 'All Users':
            for u in all_users:
                all_receipents.append(u['email'])
        if request.form['receiver'] == 'All HODs':
            for hod in all_hods:
                all_receipents.append(hod['email'])
        if request.form['receiver'] == 'All Mentors':
            for mentor in all_mentors:
                all_receipents.append(mentor['email'])
        if request.form['receiver'] == 'All Mentees':
            for mentee in all_mentees:
                all_receipents.append(mentee['email'])
        
        msg = Message('MM: '+session['fname']+' - '+request.form['subject'], sender=session['email'], recipients=all_receipents)
        msg_string = request.form['message']
        msg.body = msg_string
        msg.html = msg.body
        mail.send(msg)
        flash('Mail Sent Successfully','success')
            
    return render_template("send_mail.html")


@app.route('/dashboard',methods = ['POST','GET'])
@is_logged_in
def dashboard():
    users = mongo.db.users
    mentors1 = users.find({'type':'mentor'})
    mentee_count = {}  
    if session['type'] == 'mentor':
        for men in mentors1:
            _count = users.count_documents({'type':"mentee",'mentor_email':men['email']})
            mentee_count.__setitem__(men['email'], _count) 
        mentor = users.find_one({'email':session['email']})
        all_mentees = users.find({'type':'mentee','mentor_email':session['email']})
        return render_template("dashboard.html",all_mentees = all_mentees, mentee_count=mentee_count)
    
    elif session['type'] == 'hod':  
        for men in mentors1:
            _count = users.count_documents({'type':"mentee",'mentor_email':men['email']})
            mentee_count.__setitem__(men['email'], _count) 
        all_mentors = users.find({'type':'mentor','branch':session['branch']})
        all_mentees = users.find({'type':'mentee','branch':session['branch']})
        return render_template("dashboard.html",all_mentees = all_mentees,all_mentors = all_mentors,mentee_count=mentee_count)

    elif session['type'] == 'admin':
        for men in mentors1:
            _count = users.count_documents({'type':"mentee",'mentor_email':men['email']})
            mentee_count.__setitem__(men['email'], _count) 
        all_mentors = users.find({'type':'mentor'})
        all_mentees = users.find({'type':'mentee'})
        return render_template("dashboard.html",all_mentees = all_mentees,all_mentors = all_mentors,mentee_count=mentee_count)

@app.route('/meeting_request',methods = ['POST','GET'])
@is_logged_in
@is_mentor
def meeting_request():
    users = mongo.db.users
    meetings = mongo.db.meetings
    all_meetings = meetings.find()
    mentor = users.find_one({'email':session['email'],'type':'mentor'})
    if request.method == 'POST':
        users.update_one({'email':session['email']},{'$push':{'meeting_request':{'subject':request.form['subject'],
        'time':request.form['time'],'venue':request.form['venue'],'date':request.form['date'],
        'done':'no'}}},upsert=True)
        meeting_id = request.form['subject']+str(randint(1111,9999))
        meetings.insert_one({'mentor_email':session['email'],'meeting_id':meeting_id,'subject':request.form['subject'],'time':request.form['time'],'venue':request.form['venue'],'date':request.form['date'],'done':'no'})

        #mentor1 = users.find_one({'email':session['email'],'type':'mentor'})
        #for bat in mentor1['batch']:
        #    print(bat)
        mentees = users.find({'mentor_email':session['email']})
        
        all_receipents = []

        for mentee in mentees:
            all_receipents.append(mentee['email'])

        

        msg = Message('MM: '+session['fname']+' - Meeting regarding '+request.form['subject'], sender=session['email'], recipients=all_receipents)
        msg_string = '<h3>Subject :</h3>'+request.form['subject'] + '<br><h3>Time :</h3>' + request.form['time'] + '<br><h3>Date :</h3>' + request.form['date'] + '<br><h3>Venue :</h3>' + request.form['venue']
        msg.body = msg_string
        msg.html = msg.body
        mail.send(msg)

        flash('Meeting request generated successfully','success')
    return render_template("meeting_request.html",mentor = mentor, all_meetings = all_meetings)


@app.route('/delete_meeting/<id>', methods=['GET', 'POST'])
@is_logged_in
@is_mentor
def delete_meeting(id):
    meetings = mongo.db.meetings
    users = mongo.db.users
    #update({'type':'mentee','mentor_email':'@.com'},{"$pull":{"meetings":{"meeting_id":"779120"}}},{multi:true})
    users.update({'type':'mentee','mentor_email':session['email']},{'$pull':{"meetings":{'meeting_id':id}}},multi=True)
    meetings.delete_one({'meeting_id':id})
    flash('Meeting request deleted successfully','success')
    return redirect(url_for('meeting_request'))


@app.route('/delete/mentee/<email>')
def delete(mentee):

    
    return redirect(request.url)

@app.route('/meetings',methods = ['POST','GET'])
@is_logged_in
def meetings():
    meetings = mongo.db.meetings
    all_meetings = meetings.find({'mentor_email':session['email']}).sort('date',-1)
    return render_template("meetings.html",meetings = all_meetings)

@app.route('/meeting/<id>',methods = ['GET','POST'])
@is_logged_in
@is_mentor
def meeting(id):
    users = mongo.db.users
    meetings = mongo.db.meetings
    mentor = users.find_one({'email':session['email']})
    mentees = users.find({'type':'mentee','mentor_email':session['email']})
    meeting = meetings.find_one({'meeting_id':id})
    current_mentee = users.find_one({'email':request.args.get('email')})
    remark1 = []
    x = datetime.now()
    # x=datetime.datetime(2009, 10, 5, 18, 00)
    if current_mentee is not None:
        if 'meetings' in current_mentee:
            remark = users.find_one({'email':request.args.get('email'),'meetings.meeting_id':id},{ 'meetings': { '$elemMatch': {'meeting_id': id } } })
            if remark is not None:
                print(remark['meetings'][0]['remark'])
                remark1= remark['meetings'][0]
                print(remark1)
    if meeting['mentor_email'] == session['email']:
        if request.method == 'POST':
            meet = users.find_one({'email':request.args.get('email'),'meetings.meeting_id':id})
            #meeting entry for first time
            if meet is None:
                users.update_one({'email':request.args.get('email')},{'$push':{'meetings':{'mentor_email':session['email'],'meeting_id':meeting["meeting_id"],'subject':meeting['subject'],'time':meeting['time'],'venue':meeting['venue'],'date':meeting['date'],'done':'yes','remark':request.form['remark'],'last_updated_on':x}}})
                print('done')
            else: #mentee alreay appeared for meeting
                
                users.update_one({'email':request.args.get('email'),'meetings.meeting_id':id},{'$set':{'meetings.$.remark':request.form['remark']}})
                users.update_one({'email':request.args.get('email'),'meetings.meeting_id':id},{'$set':{'meetings.$.last_updated_on':x}})
                print('done2')
                print(request.form['remark'])
            return redirect(url_for('meeting',id=id,email=request.args.get('email')))
        return render_template("meeting.html",meeting = meeting,mentees = mentees,id=id,current_mentee =current_mentee,remark1 = remark1)
    else:
        return redirect(request.url)


@app.route('/manage_subjects',methods=['POST','GET'])
@is_logged_in
@is_mentor
def manage_subjects():
    subjects = mongo.db.subjects
    all_subjects = subjects.find()
    branch = mongo.db.branches
    branches = branch.find()
    if request.method == 'POST':
        branch_found = subjects.find_one({'branch':request.form['branch'],'subjects.semester':request.form['semester']},{'subjects':1})
        if branch_found is not None:
            for semester in branch_found['subjects']:
                if request.form['subject_name'] in semester['subject']:
                    flash('This subject is already in semester '+semester['semester'],'danger')
                    return redirect(url_for('manage_subjects'))

            subjects.update_one({'branch':request.form['branch'],'subjects.semester':request.form['semester']},{'$push':{'subjects.$.subject':request.form['subject_name']}})
            print('updated')
        else:
            subjects.update_one({'branch':request.form['branch']},{'$push':{'subjects':{'semester':request.form['semester'],'subject':[request.form['subject_name']]}}})
        
            
    return render_template("manage_subjects.html",all_subjects = all_subjects, branches=branches)

@app.route('/edit_subject',methods=['POST','GET'])
@is_logged_in
@is_mentor
def edit_subject():
    users = mongo.db.users
    subjects = mongo.db.subjects
    if request.method == "POST":
        option = request.form['subjectoption']
        if option == "Edit":
            subjects.update({"subjects.semester":request.form['semester'],"subjects.subject":request.form['old_subject']},{"$pull":{"subjects.$.subject":request.form['old_subject']}})
            subjects.update({"subjects.semester":request.form['semester']},{"$push":{"subjects.$.subject":request.form['new_subject']}})
            return redirect(url_for("manage_subjects"))
        if option == "Delete":
            subjects.update({"subjects.subject":request.form['old_subject']},{"$pull":{"subjects.$.subject":request.form["old_subject"]}})
            return redirect(url_for("manage_subjects"))
    
    return redirect(url_for("manage_subjects"))
    

@app.route('/achievements', methods=['GET', 'POST'])
@is_logged_in
def achievements():
    users = mongo.db.users
    user = users.find_one({'email':session['email']})
    if request.method == 'POST':
        done = users.update_one({'email':session['email']},{'$push':{'achievements':{'achievement':request.form['achievement'],'category':request.form['category'],'date':request.form['date']}}})
        if done is not None:
            flash('Achievement added successfully','success')
            return redirect(url_for('achievements'))
        else:
            flash('Something went wrong...!!!','danger')

    return render_template("achievements.html",user = user)
    

@app.route('/view_chart/<email>', methods=['GET', 'POST'])
@is_logged_in
def view_chart(email):
    users = mongo.db.users
    user = users.find_one({'email':email})
    user1 = users.find_one({'email':email})
    achievements_count = {
        'technical':0,
        'non_technical':0,
        'sports':0,
        'cultural':0,
        'papers_published':0,
        'other':0
    }


    for achievement in user1['achievements']:

        if achievement['category'] == 'technical':
            achievements_count['technical'] +=1
        elif achievement['category'] == 'non_technical':
            achievements_count['non_technical'] +=1
        elif achievement['category'] == 'sports':
            achievements_count['sports'] +=1
        elif achievement['category'] == 'cultural':
            achievements_count['cultural'] +=1
        elif achievement['category'] == 'papers_published':
            achievements_count['papers_published'] +=1
        elif achievement['category'] == 'other':
            achievements_count['other'] +=1



    print(achievements_count['technical'])
        

    return render_template("view_chart.html",user = user,a_count = achievements_count)



@app.route('/logout')
def logout():
    if 'logged_in' in session:
        session.clear()
        flash('Successfully logged out','success')
        return redirect(url_for('login'))
    else:
        flash('You are not Logged in','secondary')
        return redirect(url_for('login'))

@app.route('/mentee_status/<email>', methods=['GET','POST'])
@is_logged_in
@is_mentor
def mentee_status(email):
    users = mongo.db.users
    if request.method == 'POST':
        if request.form['status']:
            print("email----{}".format(email))
            print(len(request.form['status']))
            users.update_one({"email":email},{"$set":{"status":request.form['status']}},upsert=True)    
        return redirect(url_for('mentee_status',email=session['email']))
    else:
        all_mentee = users.find({"mentor_email":email})
        return render_template("mentee_status.html",mentees=all_mentee)



@app.route('/delete_mentor/<email>')
@is_logged_in
@is_admin
def delete_mentor(email):
    users = mongo.db.users
    mentee_assigned = users.count_documents({'mentor_email':email})
    if mentee_assigned > 0:
        flash('Mentor has been assigned mentees','danger')
        flash('Mentor cannot be deleted','danger')
        flash('Edit the mentor info instead','danger')
        
        return redirect(url_for('add_mentor'))
    
    users.delete_one({'email':email,'type':'mentor'})
    flash('Mentor deleted successfully','danger')
    return redirect(url_for('add_mentor'))

@app.route('/delete_hod/<email>')
@is_logged_in
@is_admin
def delete_hod(email):
    users = mongo.db.users
    branches = mongo.db.branches
    branch = users.find_one({'email':email,'type':'hod'})
    
    users.delete_one({'email':email,'type':'hod'})
    branches.delete_one({'hod_email':email})

    flash('HOD deleted successfully','danger')
    return redirect(url_for('add_hod'))


@app.route('/edit_hod/<email>', methods=['GET','POST'])
@is_logged_in
@is_admin
def edit_hod(email):
    users = mongo.db.users
    old_hod = users.find_one({'email':email, 'type':'hod'})
    if request.method == 'POST':
        find_hod = users.find_one({'email':email, 'type':'hod'})
        if find_hod is not None: 
            users.update_one({'email':email},{'$set':{'fname':request.form['fname'], 'mname':request.form['mname'], 'lname':request.form['lname'], 'email':request.form['email'], 'phone':request.form['phone'], 'branch':request.form['branch']}})
            flash('HOD Updated successfully','success')
            return redirect(url_for('add_hod',email=request.form['email']))
        else:
            flash('Something Went wrong','danger')

    return render_template('edit_hod.html', old_hod=old_hod)
    
@app.route('/edit_mentor/<email>', methods=['GET','POST'])
@is_logged_in
@is_admin
def edit_mentor(email):
    users = mongo.db.users
    meetings = mongo.db.meetings
    branches = mongo.db.branches
    all_branches = branches.find()
    old_mentor = users.find_one({'email':email, 'type':'mentor'})
    if request.method == 'POST':
        find_mentor = users.find_one({'email':email, 'type':'mentor'})
        if find_mentor is not None: 
            users.update_one({'email':email},{'$set':{'fname':request.form['fname'], 'mname':request.form['mname'], 'lname':request.form['lname'], 'email':request.form['email'], 'phone':request.form['phone'], 'branch':request.form['branch']}})
            users.update({'mentor_email':email},{'$set':{'mentor_email':request.form['email']}})
            meetings.update({'mentor_email':email},{'$set':{'mentor_email':request.form['email']}})

            flash('Mentor Updated successfully','success')
            return redirect(url_for('add_mentor'))
        else:
            flash('Something Went wrong','danger')
            return redirect(url_for('add_mentor'))

    return render_template('edit_mentor.html', old_mentor=old_mentor, all_branches=all_branches)


@app.route('/show_mentees/<email>', methods=['GET','POST'])
@is_logged_in
@is_mentor
def show_mentees(email):
    users = mongo.db.users
    all_mentees = users.find({'mentor_email':email, 'type':'mentee'})
    mentor = users.find_one({'email':email, 'type':'mentor'})
    return render_template("show_mentees.html", all_mentees=all_mentees, mentor=mentor)

@app.route('/edit_meeting/<id>', methods=['GET','POST'])
@is_logged_in
@is_mentor
def edit_meeting(id):
    meetings = mongo.db.meetings
    users = mongo.db.users
    old_meeting = meetings.find_one({'mentor_email':session['email'], 'meeting_id':id})
    if request.method == "POST":
        meetings.update({'mentor_email':session['email'],'meeting_id':id},{'mentor_email':session['email'],'meeting_id':id ,'subject':request.form['subject'], 'time':request.form['time'], 'date':request.form['date'], 'venue':request.form['venue']})
        #update({'type':'mentee','mentor_email':'cristianor383@gmail.com','meetings.meeting_id':"kkk2148"},{"$set":{"meetings.$.time":"6900"}})
        if old_meeting['done'] == 'yes':
            users.update({'type':'mentee','mentor_email':session['email'],'meetings.meeting_id':id},{"$set":{"meetings.$.subject":request.form['subject'],"meetings.$.time":request.form['time'],"meetings.$.date":request.form['date'], "meetings.$.venue":request.form['venue']}}, multi=True)
        else:
            users.update({'type':'mentee','mentor_email':session['email'],'meetings.meeting_id':id},{"$set":{"meetings.$.subject":request.form['subject'],"meetings.$.time":request.form['time'],"meetings.$.date":request.form['date'], "meetings.$.venue":request.form['venue']}}, multi=True)
        return redirect(url_for("meeting_request"))
        
    return render_template("edit_meetings.html",old_meeting=old_meeting)


@app.route('/download_status/<email>', methods=['GET','POST'])
@is_logged_in
@is_admin
def download_status(email):
    file_name = "menteeStatus.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    users = mongo.db.users
    
    length = users.count_documents({'mentor_email':email, 'type':'mentee'})
    all_mentee = users.find({'mentor_email':email, 'type':'mentee'})
    mentor_name = users.find_one({'email':email, 'type':'mentor'})
    
    
    sheet.merge_cells('A1:E1') 
    sheet.cell(row = 1, column = 1).value = "Mentee Status under guidence of "+mentor_name['fname']+" "+mentor_name['lname']
    
    sheet["A2"] = "Sr.NO."
    sheet["B2"] = "Name"
    sheet["C2"] = "Status"
    i=3
    for mentee in all_mentee:
        sheet["B"+str(i)] = mentee['fname']+" "+ mentee['mname'][0]+" " +mentee["lname"]
        sheet["C"+str(i)] = mentee['status']
        i+=1
    
    
    print(str(workbook.save(filename="menteeStatus.xlsx")))
    return send_file("menteeStatus.xlsx", as_attachment=True ,cache_timeout=0)
    

@app.route("/change_mentor/<email>",methods=['POST'])
@is_logged_in
@is_only_mentee
def change_mentor(email):
    users = mongo.db.users
    get_mentor_name = users.find_one({'email':request.form['change_mentor'], 'type':'mentor'})
    mentor_name = get_mentor_name['fname']+" "+get_mentor_name['lname']
           
    if request.method == 'POST':
        users.update_one({'email':email,'type':"mentee"},{"$set": {'mentor_email':request.form['change_mentor'],'mentor':mentor_name}})
        
        session['mentor']=get_mentor_name['email']
        session['mentor_name']= mentor_name
        flash("Mentor Change Successfull!","success")
        return redirect(url_for('profile'))
    flash("Mentor didn't Change!","danger")
    return redirect(url_for('profile'))

@app.route("/backupdatabase")
@is_logged_in
@is_admin
def backupdatabase():
    year = str(datetime.now().year)
    month = str(datetime.now().month)
    day = str(datetime.now().day)
    
    stored = os.system("mongodump -h 127.0.0.1:27017 -d mm -o "+basedir+"/Backup/"+year+"/"+month+"/"+day)
    if stored == 0:
        flash("Database Backup Successfull!","success")
    else:
        flash("Database Backup Fail!","danger")

    return redirect(url_for('index'))

if __name__ == '__main__':
    app.secret_key='secret123'
    app.run(host='0.0.0.0',debug='true',port='5000')